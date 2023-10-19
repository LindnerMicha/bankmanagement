import openpyxl

fileXLSX = openpyxl.load_workbook('Datenbanken.xlsx')

sheet = fileXLSX["User_Datenbank"]

email = sheet['A2'].value


row_count = sheet.max_row
column_count = sheet.max_column

def pw_scraper():
    id = 2
    search = input("Welches Passowort wird gesucht ? : ")

    for row in sheet['B2': f'B{row_count}']:
        for col in row:
            if col.value == search:
                print(f"Das Passwort gehört zu {sheet[f'C{id}'].value} {sheet[f'D{id}'].value}")
            id += 1

pw_scraper()


"""for row in sheet['B2': f'B{row_count}']:
    for col in row:
        print(col.coordinate, col.value)"""

